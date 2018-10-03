#coding:utf-8

import openpyxl
import os
import pandas
import xlrd






def first_row_empty(workbook, sheet, start_row, column):
    ind_col = column
    ind_row = start_row
    wb = openpyxl.load_workbook(workbook)
    ws = wb.get_sheet_by_name(sheet)

    cell_to_test = ws.cell(column = ind_col, row= ind_row)

    while cell_to_test.value != None:
        ind_row += 1
        #print(cell_to_test.value)
        cell_to_test = ws.cell(column=ind_col, row=ind_row)

    return cell_to_test.row

print(first_row_empty("Exemple1.xlsx", "Patate", 2, 1))

"""
print ("wb est un objet de type {}".format(type(wb)))
list_wb_sheet = wb.get_sheet_names()
print ("wb contient les sheet suivante{}".format(list_wb_sheet))

# une_liste = ["chat", "chien", "oiseau", "marmotte", "ecureil", "vache",]
#
# for i in une_liste:
#     print(i)
i = 0

for item in list_wb_sheet:
    print("list_wb_sheet[{}] est {}".format(i, item))
    i += 1

sheet = wb.get_sheet_by_name("Carotte")
print(sheet.title)
print("sheet est un object de type {}".format(type(sheet)))



wbb =  pandas.ExcelFile("Exemple1.xlsx")
print (wbb)

sheet = wb.get_sheet_by_name("Patate")
df = pandas.read_excel(wbb, sheet)


print (df)"""
#find last value

