#coding=utf-8
import pandas as pd
import re
from openpyxl import load_workbook
from tkinter import filedialog



def check_cell_name(cell_name):
    result = re.search("^[A-Z]+[0-9]+$", cell_name.upper())
    print()

sourcefile = filedialog.askopenfile('r')
# print(sourcefile.name)
wb = load_workbook(filename=sourcefile.name, read_only=True, data_only=True)

# wb = load_workbook(filename="Static flexibility test results.xlsx", read_only=True, data_only=True)

print(", ".join(wb.sheetnames))
ws = wb[input("Enter the sheet name: ")]

# Read the cell values into a list of lists
data_rows = []
# for row in ws['A2':'g11']:
while True:
    first_cel = input("Enter the first top left cell (A1): ")
    result = re.search("^[A-Z]+[0-9]+$", first_cel.upper())
    if result != None:
        break

while True:
    last_cel = input("Enter the last bottom right cell (Z99): ")
    result = re.search("^[A-Z]+[0-9]+$", last_cel.upper())
    if result != None:
        break
print(first_cel + " : " + last_cel)
print(type(first_cel))
for row in ws[first_cel:last_cel]:
    data_cols = []
    for cell in row:
        data_cols.append(cell.value)
    data_rows.append(data_cols)




# Transform into dataframe

df = pd.DataFrame(data_rows)
# print(df)

x_size, y_size = df.shape
with open("table.txt", 'w+', encoding='utf8') as tbl:
    tbl.writelines("")
for line in range(x_size):
    row_list = []
    for cell_data in range(y_size):
        if df.iloc[line][cell_data] is None:
            row_list.append("-")
        elif str(df.iloc[line][cell_data]).find("\n"):
            original_text = str(df.iloc[line][cell_data])
            new_string = original_text.replace("\n", " ")
            row_list.append(new_string)
        else:
            row_list.append(str(df.iloc[line][cell_data]))
    if line == 0:
        text_line = "|||:" + "|:".join(row_list) + "\n"
    else:
        text_line = "||" + "|".join(row_list) + "\n"
    # text_line = text_line[:-1]
    # print(text_line)

    # row_list.append("\n")
    # print(row_list)
    # text_line = "|" + "|".join(row_list)
    with open("table.txt", 'a+', encoding='utf8') as tbl:
        tbl.writelines(text_line)
print("Finish ! File table.txt created")