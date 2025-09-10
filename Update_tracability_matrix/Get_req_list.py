from docx import Document
import os
import pandas as pd
import openpyxl
import re

class Extrace_req:
    """
        Class to handle the extraction of requirement references from DOCX files.
        """
    def __init__(self):
        pass

    old_new_doc_num = {"DES73-6894": "DEV-0044559", "DES73-6606": "DEV-0044560", "DES73-7095": "DEV-0044581",
                       "DES73-6896": "DEV-0044561", "DES73-6897": "DEV-0044562", "DES73-6985": "DEV-0044568",
                       "DES73-6881": "DEV-0044563", "DES73-6877": "DEV-0044565", "DES73-6997": "DEV-0044569",
                       "DES73-6878": "DEV-0044566", "DES73-6973": "DEV-0044564", "DES73-6984": "DEV-0044567",
                       "DES73-6895": "DEV-0044570", "DES73-6899": "DEV-0044580", "DES73-6900": "DEV-0044573",
                       "DES73-6901": "DEV-0044574", "DES73-6902": "DEV-0044575", "DES73-7100": "DEV-0044571",
                       "DES73-6904": "DEV-0044576", "DES73-": "DEV-0044", "DES73-": "DEV-0044",
                       "DES73-": "DEV-0044", "DES73-6882": "DEV-0044577", "DES73-": "DEV-0044",
                       "DES73-": "DEV-0044", "DES73-": "DEV-0044", "DES73-": "DEV-0044",
                       "DES73-": "DEV-0044", "DES73-": "DEV-0044", "DES73-": "DEV-0044",
                       "DES73-": "DEV-0044", "DES73-": "DEV-0044", "DES73-": "DEV-0044",
                       "DES73-": "DEV-0044", "DES73-": "DEV-0044", "DES73-": "DEV-0044",
                       "DES73-": "DEV-0044", "DES73-": "DEV-0044", "DES73-": "DEV-0044",
                       "DES73-": "DEV-0044", "DES73-": "DEV-0044", "DES73-": "DEV-0044",
                       "DES73-": "DEV-0044", "DES73-": "DEV-0044", "DES73-": "DEV-0044",}

    def list_folder(self, path):
        """
                Lists all file paths in the specified directory.

                Args:
                    path (str): The folder path.

                Returns:
                    list: List of full file paths.
                """
        folder_path = path  # Change this to your folder path
        file_paths = [os.path.join(folder_path, f) for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f))]
        return file_paths

    def extract_req_from_doc(self, files_list):
        """
               Extracts requirements from the first table of each DOCX file.

               Args:
                   files_list (list): List of DOCX file paths.

               Returns:
                   dict: Dictionary mapping document numbers to extracted requirements.
               """
        result = {}
        for file in files_list:
            doc = Document(file)
            tbl = doc.tables[0]
            cell = tbl.cell(2, 1)
            text = cell.text
            result[(self._extract_doc_number_from_path(file))] = self._create_array_of_req(text)
        return result

    def _extract_doc_number_from_path(self, path):
        """
                Extracts the document number from the file path.

                Args:
                    path (str): Full file path.

                Returns:
                    str: First 10 characters of the file name.
                """

        sep_path = path.split("\\")
        if sep_path[-1][0:3] == "DES":
            return self.old_new_doc_num[sep_path[-1][0:10]]
        else:
            return sep_path[-1][0:11]
        pass

    def _create_array_of_req(self, text):
        """
               Extracts requirement IDs in the format `X.XXX` using regex.

               Args:
                   text (str): Text content from the DOCX cell.

               Returns:
                   list: List of matching requirement references.
               """
        cleaned_list = re.findall(r"[0-9]{1,3}\.[0-9]{3}", text)
        return cleaned_list

    def pd_dataframe_from_dict(self, data):
        """
                Converts a dictionary to a Pandas DataFrame.

                Args:
                    data (dict): Dictionary to convert.

                Returns:
                    pd.DataFrame: Converted DataFrame.
                """
        df = pd.DataFrame.from_dict(data, orient="index").T
        return df

class Excel_manager:
    """
        Class to manage Excel operations using openpyxl.
        """
    def __init__(self):
        pass

    def open_excel(self, file, sheet):
        """
                Opens an Excel file and selects a worksheet.

                Args:
                    file (str): Excel file path.
                    sheet (str): Sheet name to open.

                Returns:
                    Worksheet object.
                """
        self.wb = openpyxl.load_workbook(file)
        self.ws = self.wb[sheet]
        return self.ws

    def read_col_extract_PRS(self):
        """
                Reads column B and extracts all values starting with 'PR_'.

                Returns:
                    list: List of requirement IDs (e.g., PR_1234).
                """
        prs_list = []
        col = self.ws['B']
        for c in col:
            try:
                if c.value[0:3] == "PR_":
                    prs_list.append(c.value)
            except TypeError:
                pass
        return prs_list

    def cell_coord(self, value, col):
        """
                Finds the cell in a column matching a specific value.

                Args:
                    value (str): Cell value to find.
                    col (str): Column letter (e.g., 'B').

                Returns:
                    Cell object if found, else None.
                """
        col = self.ws[col]
        for c in col:
            if c.value == value:
                return c

    def truncate(self, data):
        """
               Truncates the first 3 characters from a string.

               Args:
                   data (str): Input string.

               Returns:
                   str: Truncated string.
               """
        return data[3:]

    def marking(self, col_to_mark, req_list):
        """
                Marks an 'x' in the specified column for each requirement ID.

                Args:
                    col_to_mark (str): Excel column to mark.
                    req_list (list): List of requirement IDs (without 'PR_' prefix).
                """
        for req in req_list:
            req = "PR_{}".format(req)
            cell_row = self.cell_coord(req, "B")
            try:
                cell = "{}{}".format(col_to_mark, cell_row.row)
                self.ws[cell].value = "x"
            except Exception as e:
                print(f"The text requirement is: {req}")
                print(f"{e}")


    def generate_doc_id_corolation(self):
        document_list = {}
        starting_col = 8
        ending_col = 71
        collect_row = 3
        for col_id in range(starting_col, ending_col):
            document_list[self.ws.cell(collect_row, col_id).value[0:11]] = self.ws.cell(collect_row, col_id).column_letter
        return document_list


def main():
    """
        Main execution function that:
        - Maps document IDs to Excel columns.
        - Extracts requirements from DOCX files.
        - Marks corresponding cells in the Excel sheet.
        - Saves the result in a new Excel file.
        """


    o = Extrace_req()
    exl = Excel_manager()
    ws = exl.open_excel("DEV-0044585 STMN IOS Software Traceability Matrix Rev 5.xlsx", "1.0.3")
    folder_list = [r"C:\Users\u120230\temp\SDD", r"C:\Users\u120230\temp\VER", r"C:\Users\u120230\temp\VAL"]

    col_id_relation = exl.generate_doc_id_corolation()

    for folder in folder_list:
        file_list = o.list_folder(folder)
        result = o.extract_req_from_doc(file_list)
        for key, value in result.items():
            print(key)
            exl.marking(col_id_relation[key], value)
    exl.wb.save("out.xlsx")

def test():
    exl = Excel_manager()
    exl.generate_doc_id_corolation()

if __name__ == "__main__":
    main()
