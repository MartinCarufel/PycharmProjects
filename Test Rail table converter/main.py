import re
from tkinter import filedialog
from kivy.app import App
from kivy.metrics import dp
from kivy.uix.button import Button
from kivy.uix.stacklayout import StackLayout
from kivy.uix.textinput import TextInput
from kivy.uix.togglebutton import ToggleButton
from kivy.uix.widget import Widget
from openpyxl import load_workbook
import pandas as pd


class MainWidget(Widget):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        sourcefile = filedialog.askopenfile('r')

        # print(sourcefile.name)
        self.wb = load_workbook(filename=sourcefile.name, read_only=True, data_only=True)
        # self.wb = load_workbook(filename="Book1.xlsx", read_only=True, data_only=True)

        # wb = load_workbook(filename="Static flexibility test results.xlsx", read_only=True, data_only=True)

        # print(", ".join(self.wb.sheetnames))


class StackLayoutEx(StackLayout, MainWidget):

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.orientation = 'lr-tb'
        self.start_cell_input = None
        self.end_cell_input = None

        # Padding = gauche, haut, droit, bas
        self.padding = dp(10)
        self.spacing = dp(2)

        # for i in range(100):
        #     b = Button(text=str(i+1), size_hint=(None, None), size=(dp(60), dp(60)))
        #     self.add_widget(b)

        for i in self.wb.sheetnames:
            b = ToggleButton(text=i, size_hint=(None, None), size=(dp(260), dp(60)), group="sheet")
            # b = Button(text=i, size_hint=(None, None),)
            self.add_widget(b)
            b.bind(on_press=self.callback)

        self.start_cell_input = TextInput(size_hint=(None, None), size=(dp(60), dp(60)), multiline=False)
        self.start_cell_input.bind(text=self.on_text)
        self.add_widget(self.start_cell_input)

        self.end_cell_input = TextInput(size_hint=(None, None), size=(dp(60), dp(60)), multiline=False)
        self.end_cell_input.bind(text=self.on_text)
        self.add_widget(self.end_cell_input)

        validate = Button(text="Convert", size_hint=(None, None), size=(dp(260), dp(60)))
        validate.bind(on_press=self.validate)
        self.add_widget(validate)

    def on_text(self, value, other):
        pass
        # print(value.text)

    def validate(self, event):
        # print(self.ws)
        # self.wb == load_workbook(filename="Book1.xlsx", read_only=True, data_only=True)
        ws = self.wb[self.ws]
        data_rows = []
        # for row in ws['A2':'g11']:
        first_cel = self.start_cell_input.text.upper()
        last_cel = self.end_cell_input.text.upper()

        # print(first_cel + " : " + last_cel)
        # print(type(first_cel))
        for row in ws[first_cel:last_cel]:
            data_cols = []
            for cell in row:
                data_cols.append(cell.value)
            data_rows.append(data_cols)

        # Transform into dataframe

        df = pd.DataFrame(data_rows)
        print(df)

        x_size, y_size = df.shape
        with open("table.txt", 'w+', encoding='utf8') as tbl:
            tbl.writelines("")
        for line in range(x_size):
            row_list = []
            for cell_data in range(y_size):
                if df.iloc[line][cell_data] is None:
                    row_list.append("-")
                elif str(df.iloc[line][cell_data]).find("\n"):
                    original_text = str(df.iloc[line][cell_data])
                    new_string = original_text.replace("\n", " ")
                    row_list.append(new_string)
                else:
                    row_list.append(str(df.iloc[line][cell_data]))
            if line == 0:
                text_line = "|||:" + "|:".join(row_list) + "\n"
            else:
                text_line = "||" + "|".join(row_list) + "\n"
            # text_line = text_line[:-1]
            # print(text_line)

            # row_list.append("\n")
            # print(row_list)
            # text_line = "|" + "|".join(row_list)
            with open("table.txt", 'a+', encoding='utf8') as tbl:
                tbl.writelines(text_line)

        print("Finish ! File table.txt created")
        # self.wb.close()

    def callback(self, event):
        self.ws = event.text
        test = self.wb[self.ws]
        last_empty_row = len(list(test.rows))
        l = list(self.wb[self.ws].iter_rows())
        line_num = 1
        for i in l:
            print(line_num, end=' | ')
            for cel in i:

                print(cel.value, end=' | ')
            line_num += 1
            print()
        # newbd = pd.DataFrame(l)
        # print(newbd)

        # print(event.text)










class ExcelSheetSeletor(App):
    pass

ExcelSheetSeletor().run()
